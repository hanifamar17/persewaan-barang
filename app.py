import tempfile
from flask import Flask, request, redirect, send_file, url_for, render_template, flash, jsonify, abort, Response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import User
from db import get_db_connection
from werkzeug.security import check_password_hash, generate_password_hash
import os
import uuid
from flask_wtf.csrf import CSRFProtect, CSRFError
from datetime import datetime, date
import traceback
import pdfkit
import base64
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import csv


app = Flask(__name__)
secret_key = os.urandom(24)
app.secret_key = secret_key
csrf = CSRFProtect(app)


login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

# PDFKit configuration using wkhtmltopdf
WKHTMLTOPDF_PATH= r'E:\\App-Development\\1-Tools-Library-Environment\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
pdfkit_config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)

## OTHERS
#--- Format rupiah --- 
@app.template_filter('format_rupiah')
def format_rupiah(amount):
    try:
        amount = float(amount)
        return f"Rp{amount:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return amount 

# --- Format tanggal Indonesia ---
def format_tanggal_indonesia(tanggal):
    bulan_id = {
        'January': 'Januari',
        'February': 'Februari',
        'March': 'Maret',
        'April': 'April',
        'May': 'Mei',
        'June': 'Juni',
        'July': 'Juli',
        'August': 'Agustus',
        'September': 'September',
        'October': 'Oktober',
        'November': 'November',
        'December': 'Desember'
    }
    if isinstance(tanggal, str):
        tanggal = datetime.strptime(tanggal, "%Y-%m-%d")  # Sesuaikan format aslinya

    bulan = bulan_id[tanggal.strftime("%B")]
    return f"{tanggal.day} {bulan} {tanggal.year}"

app.jinja_env.filters['tanggal_id'] = format_tanggal_indonesia

# --- Daftar bulan dalam bahasa Indonesia ---
bulan_id = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

# --- Load user by ID for session ---
@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
    user_data = cursor.fetchone()
    conn.close()
    if user_data:
        return User(user_data['user_id'], user_data['name'], user_data['username'], user_data['password'], user_data['role'])
    return None

@app.route('/', methods=['GET', 'POST'])
def index():
    return redirect(url_for('login'))

# --- Route: Login ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password_input = request.form['password']

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
        user_data = cursor.fetchone()
        conn.close()

        if user_data and check_password_hash(user_data['password'], password_input):
            user = User(user_data['user_id'], user_data['name'], user_data['username'], user_data['password'], user_data['role'])
            login_user(user)
            flash('Login successful', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'danger')

    return render_template('login.html')

# --- Route: Logout ---
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- Route: Dashboard (akses berdasarkan role) ---
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    # 1. Total transaksi
    cursor.execute("SELECT COUNT(*) FROM transactions")
    total_transaksi = cursor.fetchone()[0]

    # 2. Total produk disewa
    cursor.execute("SELECT SUM(product_qty) FROM transaction_details")
    total_produk_diseswa = cursor.fetchone()[0] or 0

    # 3. produk yang sering disewa
    cursor.execute("""
        SELECT p.name AS product_name, c.name AS category_name, SUM(td.product_qty) as total
        FROM transaction_details td
        JOIN products p ON td.product_id = p.product_id
        JOIN categories c ON p.category_id = c.category_id
        GROUP BY p.name, c.name
        ORDER BY total DESC
        LIMIT 3
    """)
    top_produk = cursor.fetchall()  # list of tuples: (product_name, total)

    # 4. Total produk yang belum kembali
    cursor.execute("""
        SELECT SUM(td.product_qty)
        FROM transaction_details td
        JOIN transactions t ON td.transaction_id = t.transaction_id
        WHERE t.status_pembayaran != 'lunas'
    """)
    total_belum_kembali = cursor.fetchone()[0] or 0

    # 5. Total pendapatan bulan ini
    now = datetime.now()
    bulan_ini = now.month
    tahun_ini = now.year

    cursor.execute("""
        SELECT SUM(td.product_qty * p.harga_sewa * t.lama_sewa) AS total_pendapatan
        FROM transaction_details td
        JOIN transactions t ON td.transaction_id = t.transaction_id
        JOIN products p ON td.product_id = p.product_id
        WHERE MONTH(t.tanggal_nota) = %s AND YEAR(t.tanggal_nota) = %s
    """, (bulan_ini, tahun_ini))
    total_per_bulan = cursor.fetchone()[0] or 0

    cursor.close()
    conn.close()

    now = datetime.now()
    bulan = bulan_id[now.month - 1]
    return render_template(
        'index.html',
        user=current_user,
        total_transaksi=total_transaksi,
        total_produk_diseswa=total_produk_diseswa,
        top_produk=top_produk,
        total_belum_kembali=total_belum_kembali,
        total_per_bulan=total_per_bulan,
        bulan=bulan
    )

## MODULE MANAJEMEN USER
#--- Route: Halaman user --- 
@app.route('/user')
@login_required
def user():
    if current_user.role not in ['superadmin', 'admin']:
        return "Access denied", 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users")
    users = cursor.fetchall()
    conn.close()

    return render_template('admin/user.html', users=users, user=current_user)

# --- Route: Tambah User Baru (Hanya untuk admin) ---
@app.route('/add_user', methods=['GET', 'POST'])
@login_required
def add_user():
    if current_user.role not in ['superadmin', 'admin']:
        return "Access denied", 403

    if request.method == 'POST':
        try:
            name = request.form['name']
            username = request.form['username']
            password = generate_password_hash(request.form['password'])
            email = request.form.get('email', "-")
            phone_number = request.form.get('phone_number', "-")
            address = request.form.get('address', "-")
            role = request.form['role']  # admin or customer service

            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO users (name, username, password, email, phone_number, address, role)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (name, username, password, email, phone_number, address, role))
            conn.commit()
            conn.close()

            return jsonify(status="success", message="User berhasil ditambahkan")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal menambahkan user: {str(e)}"), 500

    return render_template('admin/add_user.html', user=current_user)

#--- Route: Hapus user --- 
@app.route('/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if current_user.role not in ['superadmin', 'admin']:
        return jsonify(status='error', message='Access denied'), 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Cek apakah user yang akan dihapus adalah superadmin
    cursor.execute("SELECT role FROM users WHERE user_id = %s", (user_id,))
    user = cursor.fetchone()

    if not user:
        conn.close()
        return jsonify(status='error', message='User not found'), 404

    if user['role'] == 'superadmin':
        conn.close()
        return jsonify(status='error', message='Cannot delete superadmin'), 403

    try:
        cursor.execute("DELETE FROM users WHERE user_id = %s", (user_id,))
        conn.commit()
        conn.close()
        return jsonify(status='success', message='User deleted successfully')
    except Exception as e:
        conn.close()
        return jsonify(status='error', message=f'Failed to delete user: {str(e)}'), 500


@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if current_user.role not in ['superadmin', 'admin']:
        return "Access denied", 403

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data user yang akan diedit
    cursor.execute("SELECT * FROM users WHERE user_id = %s", (user_id,))
    user_data = cursor.fetchone()

    if not user_data:
        flash('User not found', 'danger')
        conn.close()
        return redirect(url_for('user'))

    if request.method == 'POST':
        try:
            name = request.form['name']
            username = request.form['username']
            new_password = request.form.get('password')  # bisa kosong

            email = request.form.get('email', "-")
            phone_number = request.form.get('phone_number', "-")
            address = request.form.get('address', "-")

            # Ambil data user lama dulu untuk password dan role
            cursor.execute("SELECT * FROM users WHERE user_id=%s", (user_id,))
            user_data = cursor.fetchone()
            if not user_data:
                return jsonify(status="error", message="User tidak ditemukan"), 404

            # Role hanya boleh diedit jika user yang diedit bukan superadmin
            if user_data['role'] != 'superadmin':
                role = request.form['role']
            else:
                role = user_data['role']  # tetap gunakan role lama

            if new_password and new_password.strip() != "":
                # Jika user isi password baru, hash dan update
                password_hashed = generate_password_hash(new_password)
            else:
                # Jika kosong, gunakan password lama dari DB
                password_hashed = user_data['password']

            cursor.execute("""
                UPDATE users
                SET name=%s, username=%s, password=%s, email=%s, phone_number=%s, address=%s, role=%s
                WHERE user_id=%s
            """, (name, username, password_hashed, email, phone_number, address, role, user_id))

            conn.commit()
            conn.close()

            return jsonify(status="success", message="User berhasil diperbarui")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal memperbarui user: {str(e)}"), 500


    conn.close()
    return render_template('admin/edit_user.html', user=user_data)

#--- Route: Halaman Pelanggan --- 
@app.route('/customers')
@login_required
def customers():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM customers")
    customers = cursor.fetchall()
    conn.close()

    return render_template('customers/customer.html', users=customers, user=current_user)

# --- Route: Tambah Pelanggan Baru ---
@app.route('/add_customer', methods=['GET', 'POST'])
@login_required
def add_customer():
    if request.method == 'POST':
        try:
            name = request.form['name']
            phone_number = request.form.get('phone_number', "-")
            address = request.form.get('address', "-")

            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO customers (name, phone_number, address)
                VALUES (%s, %s, %s)
            """, (name, phone_number, address))
            conn.commit()
            conn.close()

            return jsonify(status="success", message="Pelanggan berhasil ditambahkan")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal menambahkan pelanggan: {str(e)}"), 500

    return render_template('customers/add_customer.html', user=current_user)

#--- Route: Hapus pelanggan --- 
@app.route('/delete_customer/<int:customer_id>', methods=['POST'])
@login_required
def delete_customer(customer_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("DELETE FROM customers WHERE customer_id = %s", (customer_id,))
        conn.commit()
        conn.close()
        return jsonify(status='success', message='Customer deleted successfully')
    except Exception as e:
        conn.close()
        return jsonify(status='error', message=f'Failed to delete customer: {str(e)}'), 500
    
#--- Route: Edit pelanggan --- 
@app.route('/edit_customer/<int:customer_id>', methods=['GET', 'POST'])
@login_required
def edit_customer(customer_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil data user yang akan diedit
    cursor.execute("SELECT * FROM customers WHERE customer_id = %s", (customer_id,))
    customer_data = cursor.fetchone()

    if not customer_data:
        flash('Customer not found', 'danger')
        conn.close()
        return redirect(url_for('customers'))

    if request.method == 'POST':
        try:
            name = request.form['name']
            phone_number = request.form.get('phone_number', "-")
            address = request.form.get('address', "-")

            # Ambil data user lama dulu untuk password dan role
            cursor.execute("SELECT * FROM customers WHERE customer_id=%s", (customer_id,))
            customer_data = cursor.fetchone()
            if not customer_data:
                return jsonify(status="error", message="Pelanggan tidak ditemukan"), 404

            cursor.execute("""
                UPDATE customers
                SET name=%s, phone_number=%s, address=%s
                WHERE customer_id=%s
            """, (name, phone_number, address, customer_id))

            conn.commit()
            conn.close()

            return jsonify(status="success", message="Pelanggan berhasil diperbarui")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal memperbarui pelanggan: {str(e)}"), 500


    conn.close()
    return render_template('customers/edit_customer.html', user=customer_data)


## MODULE MANAJEMEN PRODUK
#--- Route: Halaman Kategori --- 
@app.route('/categories')
@login_required
def categories():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM categories")
    categories = cursor.fetchall()

    conn.close()

    return render_template('products/category.html', categories=categories, user=current_user)

# --- Route: Tambah category ---
@app.route('/add_category', methods=['GET', 'POST'])
@login_required
def add_category():
    if request.method == 'POST':
        try:
            name = request.form['name']
            description = request.form.get('description',"-")

            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO categories (name, description)
                VALUES (%s, %s)
            """, (name, description))
            conn.commit()
            conn.close()

            return jsonify(status="success", message="Kategori berhasil ditambahkan")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal menambahkan kategori: {str(e)}"), 500

    return render_template('products/add_category.html', user=current_user)

#--- Route: Edit category --- 
@app.route('/edit_category/<int:category_id>', methods=['POST', 'GET'])
def edit_category(category_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT * FROM categories WHERE category_id = %s", (category_id,))
    category_data = cursor.fetchone()

    if not category_data:
        conn.close()
        return jsonify(status="error", message="Category tidak ditemukan"), 404

    try:
        name = request.form['name']
        description = request.form.get('description', "-")

        cursor.execute("""
            UPDATE categories
            SET name=%s, description=%s
            WHERE category_id=%s
        """, (name, description, category_id))

        conn.commit()
        conn.close()

        return jsonify(status="success", message="Kategori berhasil diperbarui")
    except Exception as e:
        conn.close()
        return jsonify(status="error", message=f"Gagal memperbarui kategori: {str(e)}"), 500

#--- Route: Hapus category --- 
@app.route('/delete_category/<int:category_id>', methods=['POST'])
@login_required
def delete_category(category_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("DELETE FROM categories WHERE category_id = %s", (category_id,))
        conn.commit()
        conn.close()
        return jsonify(status='success', message='Category deleted successfully')
    except Exception as e:
        conn.close()
        return jsonify(status='error', message=f'Failed to delete category: {str(e)}'), 500
    
#--- Route: Halaman produk --- 
@app.route('/products')
@login_required
def products():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil produk dan kategori dengan JOIN
    cursor.execute("""
        SELECT p.*, c.name AS category_name
        FROM products p
        JOIN categories c ON p.category_id = c.category_id
    """)
    products = cursor.fetchall()

    cursor.execute("SELECT category_id, name FROM categories")
    categories = cursor.fetchall()

    conn.close()

    # filter format rupiah
    app.jinja_env.filters['rupiah'] = format_rupiah

    return render_template('products/product.html', products=products, categories=categories, user=current_user)

# --- Route: Tambah product ---
@app.route('/add_product', methods=['GET', 'POST'])
def add_product():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            product_id = request.form['product_id']
            category_id = request.form['category_id']
            name = request.form['name']
            qty = int(request.form['qty'])
            harga_sewa = float(request.form['harga_sewa'])
            description = request.form.get('description', '-')
            status = request.form.get('status', 'tersedia')

            cursor.execute("""
                INSERT INTO products (product_id, category_id, name, qty, harga_sewa, description, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (product_id, category_id, name, qty, harga_sewa, description, status))
            conn.commit()
            return jsonify(status="success", message="Produk berhasil ditambahkan")
        except Exception as e:
            return jsonify(status="error", message=f"Gagal menambahkan produk: {str(e)}"), 500
        finally:
            conn.close()

    # untuk GET request: ambil daftar kategori
    cursor.execute("SELECT category_id, name FROM categories")
    categories = cursor.fetchall()
    conn.close()

    return render_template('products/add_product.html', categories=categories, user=current_user)

#--- Route: Edit product --- 
@app.route('/edit_product/<int:id>', methods=['POST'])
def edit_product(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        product_id = request.form['product_id']
        category_id = request.form['category_id']
        name = request.form['name']
        qty = int(request.form['qty'])
        harga_sewa = float(request.form['harga_sewa'])
        description = request.form.get('description', '-')

        cursor.execute("""
            UPDATE products
            SET product_id=%s, category_id=%s, name=%s, qty=%s, harga_sewa=%s, description=%s
            WHERE id=%s
        """, (product_id, category_id, name, qty, harga_sewa, description, id))

        conn.commit()
        return jsonify(status="success", message="Produk berhasil diperbarui")
    except Exception as e:
        return jsonify(status="error", message=str(e)), 500
    finally:
        conn.close()
    
#--- Route: Hapus product --- 
@app.route('/delete_product/<int:id>', methods=['POST'])
@login_required
def delete_product(id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("DELETE FROM products WHERE id = %s", (id,))
        conn.commit()
        conn.close()
        return jsonify(status='success', message='Product deleted successfully')
    except Exception as e:
        conn.close()
        return jsonify(status='error', message=f'Failed to delete product: {str(e)}'), 500


## MODULE TRANSAKSI
#--- Route: Halaman transaksi --- 
@app.route('/sewa')
@login_required
def sewa():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM products")
    products = cursor.fetchall()

    cursor.execute("SELECT * FROM customers")
    customers = cursor.fetchall()

    conn.close()

    # Generate nomor nota otomatis
    no_nota = f"NOTA-{uuid.uuid4().hex[:8].upper()}"

    return render_template('transactions/transaction.html', 
                           products=products, customers=customers, user=current_user,
                        no_nota=no_nota,
                           today=date.today())

#--- Route: cari ketersediaan produk ---
@app.route('/search-products', methods=['POST'])
def search_products():
    data = request.get_json()
    search_term = data.get('search_term', '')
    tanggal_sewa = data.get('tanggal_sewa')
    tanggal_kembali = data.get('tanggal_kembali')
    
    if not tanggal_sewa or not tanggal_kembali:
        return jsonify({'error': 'Tanggal sewa dan kembali harus diisi'}), 400
    
    # Validasi tanggal
    try:
        sewa_date = datetime.strptime(tanggal_sewa, '%Y-%m-%d').date()
        kembali_date = datetime.strptime(tanggal_kembali, '%Y-%m-%d').date()
        
        if sewa_date >= kembali_date:
            return jsonify({'error': 'Tanggal kembali harus setelah tanggal sewa'}), 400
            
    except ValueError:
        return jsonify({'error': 'Format tanggal tidak valid'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT 
            p.product_id, 
            p.name, 
            p.harga_sewa, 
            p.qty,
            COALESCE(SUM(
                CASE 
                    WHEN t.status_pengembalian = 'belum' 
                    AND t.tanggal_sewa < %s 
                    AND t.tanggal_kembali > %s 
                    THEN td.product_qty 
                    ELSE 0 
                END
            ), 0) as qty_disewa
        FROM products p
        LEFT JOIN transaction_details td ON p.product_id = td.product_id
        LEFT JOIN transactions t ON td.transaction_id = t.transaction_id
        WHERE (p.name LIKE %s OR p.product_id LIKE %s)
        GROUP BY p.product_id, p.name, p.harga_sewa, p.qty
        HAVING (p.qty - qty_disewa) > 0
        ORDER BY p.name
    """

    search_pattern = f'%{search_term}%'
    cursor.execute(query, (tanggal_kembali, tanggal_sewa, search_pattern, search_pattern))

    products = cursor.fetchall()

    # Hitung stok tersedia
    for product in products:
        product['stok_tersedia'] = product['qty'] - product['qty_disewa']

    cursor.close()
    conn.close()

    
    return jsonify(products)

#--- Route: Simpan transaksi ---
@app.route("/simpan_transaksi", methods=["POST"])
def simpan_transaksi():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Ambil data form utama
        customer_name = request.form.get("customer_name")
        customer_phone = request.form.get("customer_phone")
        customer_address = request.form.get("customer_address", "-") 
        user_id = request.form.get("user_id")
        no_nota = request.form.get("no_nota")
        tanggal_nota = request.form.get("tanggal_nota")
        tanggal_sewa = request.form.get("tanggal_sewa")
        tanggal_kembali = request.form.get("tanggal_kembali")
        jaminan = request.form.get("jaminan", "-")
        note = request.form.get("note", "-")
        status_pembayaran = request.form.get("status_pembayaran")

        # Cari customer_id, atau buat baru jika tidak ada
        cursor.execute("SELECT customer_id FROM customers WHERE name = %s", (customer_name,))
        customer = cursor.fetchone()
        if customer:
            customer_id = customer["customer_id"]
        else:
            cursor.execute("INSERT INTO customers (name, phone_number, address) VALUES (%s, %s, %s)", (customer_name, customer_phone, customer_address))
            conn.commit()
            customer_id = cursor.lastrowid

        # Hitung total
        lama_sewa_list = request.form.get("lama_sewa[]") or "1"
        qty_list = request.form.getlist("qty[]")
        harga_list = request.form.getlist("harga_sewa[]")

        # Validasi & konversi
        try:
            lama_sewa = int(lama_sewa_list)
        except ValueError:
            lama_sewa = 1  # fallback default jika input tidak valid

        # Hitung total dari subtotal tiap produk
        total = sum(
            int(qty) * float(harga) * lama_sewa
            for qty, harga in zip(qty_list, harga_list)
        )

        # Simpan ke tabel transactions
        cursor.execute("""
            INSERT INTO transactions (customer_id, user_id, no_nota, tanggal_nota, tanggal_sewa, tanggal_kembali, lama_sewa, status_pembayaran, jaminan, note, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (customer_id, user_id, no_nota, tanggal_nota, tanggal_sewa, tanggal_kembali, lama_sewa_list, status_pembayaran, jaminan, note, total))
        conn.commit()
        transaction_id = cursor.lastrowid

        # Simpan ke tabel transaction_details
        product_ids = request.form.getlist("product_id[]")
        for product_id, qty, harga in zip(product_ids, qty_list, harga_list):
            cursor.execute("""
                INSERT INTO transaction_details (transaction_id, product_id, product_qty, harga_sewa)
                VALUES (%s, %s, %s, %s)
            """, (transaction_id, product_id, qty, harga))

        conn.commit()
        conn.close()

        #return redirect(url_for("sewa"))
        return jsonify({
            'status': 'success',
            'message': 'Transaksi berhasil disimpan',
            'redirect_url': url_for('detail_transaksi', transaction_id=transaction_id)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


## MODULE RIWAYAT TRANSAKSI
#--- Route: Halaman riwayat transaksi --- 
@app.route('/riwayat_transaksi')
@login_required
def riwayat_transaksi():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Ambil parameter rentang waktu dari form (default: 30 hari terakhir)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        cursor.execute("""
             SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.tanggal_kembali,
                   c.name AS customer_name, t.status_pembayaran, t.status_pengembalian, t.total,
                   GROUP_CONCAT(CONCAT(p.product_id, ' - ', p.name) SEPARATOR '\n') AS produk
            FROM transactions t
            JOIN customers c ON t.customer_id = c.customer_id
            LEFT JOIN transaction_details td ON t.transaction_id = td.transaction_id
            LEFT JOIN products p ON td.product_id = p.product_id
            GROUP BY t.transaction_id
            ORDER BY t.tanggal_nota DESC
            LIMIT 50
        """)
    else:
        cursor.execute("""
            SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.tanggal_kembali,
                   c.name AS customer_name, t.status_pembayaran, t.total,
                   GROUP_CONCAT(CONCAT(p.product_id, ' - ', p.name) SEPARATOR '\n') AS produk
            FROM transactions t
            JOIN customers c ON t.customer_id = c.customer_id
            LEFT JOIN transaction_details td ON t.transaction_id = td.transaction_id
            LEFT JOIN products p ON td.product_id = p.product_id
            WHERE t.tanggal_nota BETWEEN %s AND %s
            GROUP BY t.transaction_id
            ORDER BY t.tanggal_nota DESC
        """, (start_date, end_date))

    transactions = cursor.fetchall()
    conn.close()

    return render_template('transactions/riwayat_transaksi.html', transactions=transactions, start_date=start_date, end_date=end_date, user=current_user)

#--- Route: Halaman detai riwayat transaksi --- 
@app.route('/transaksi/detail/<int:transaction_id>')
@login_required
def detail_transaksi(transaction_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Ambil transaksi utama
        cursor.execute("""
            SELECT 
                t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.tanggal_kembali,
                t.lama_sewa, t.status_pembayaran, t.status_pengembalian, t.total, t.jaminan, t.note,
                c.name AS customer_name, c.phone_number AS customer_phone, c.address AS customer_address,
                u.name AS user_name
            FROM transactions t
            JOIN customers c ON t.customer_id = c.customer_id
            JOIN users u ON t.user_id = u.user_id
            WHERE t.transaction_id = %s
        """, (transaction_id,))
        transaction = cursor.fetchone()
        if not transaction:
            abort(404)

        # Ambil detail produk
        cursor.execute("""
            SELECT 
                td.product_qty, td.harga_sewa, p.name AS product_name, t.lama_sewa, p.product_id,
                (td.product_qty * td.harga_sewa * t.lama_sewa) AS subtotal
            FROM transaction_details td
            JOIN products p ON td.product_id = p.product_id
            JOIN transactions t ON td.transaction_id = t.transaction_id
            WHERE td.transaction_id = %s
        """, (transaction_id,))
        details = cursor.fetchall()

        # Konversi ke float agar aman diformat dengan filter rupiah
        for item in details:
            item['harga_sewa'] = float(item.get('harga_sewa', 0))
            item['subtotal'] = float(item.get('subtotal', 0))

        transaction['total'] = float(transaction.get('total', 0))

        cursor.close()
        conn.close()

        app.jinja_env.filters['rupiah'] = format_rupiah
        
        return render_template(
            'transactions/detail_transaksi.html',
            transaction=transaction,
            details=details,
            user=current_user
        )

    except Exception as e:
        print(f"[ERROR] Gagal mengambil detail transaksi: {e}")
        abort(500)

#--- Route: Halaman edit transaksi --- 
@app.route('/transaksi/edit/<int:transaction_id>', methods=['GET', 'POST'])
@login_required
def edit_transaksi(transaction_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            # Ambil data form utama
            customer_name = request.form.get("customer_name")
            customer_phone = request.form.get("customer_phone")
            customer_address = request.form.get("customer_address", "-") 
            user_id = request.form.get("user_id")
            no_nota = request.form.get("no_nota")
            tanggal_nota = request.form.get("tanggal_nota")
            tanggal_sewa = request.form.get("tanggal_sewa")
            tanggal_kembali = request.form.get("tanggal_kembali")
            jaminan = request.form.get("jaminan", "-")
            note = request.form.get("note", "-")
            status_pembayaran = request.form.get("status_pembayaran")
            status_penegembalian = request.form.get("status_pengembalian")

            # Cari customer_id, atau buat baru jika tidak ada
            cursor.execute("SELECT customer_id FROM customers WHERE name = %s", (customer_name,))
            customer = cursor.fetchone()
            if customer:
                customer_id = customer["customer_id"]
            else:
                cursor.execute("INSERT INTO customers (name, phone_number, address) VALUES (%s, %s, %s)", (customer_name, customer_phone, customer_address))
                conn.commit()
                customer_id = cursor.lastrowid

            # Hitung total
            lama_sewa_list = request.form.get("lama_sewa[]") or "1"
            qty_list = request.form.getlist("qty[]")
            harga_list = request.form.getlist("harga_sewa[]")

            # Validasi & konversi
            try:
                lama_sewa = int(lama_sewa_list)
            except ValueError:
                lama_sewa = 1  # fallback default jika input tidak valid

            # Hitung total dari subtotal tiap produk
            total = sum(
                int(qty) * float(harga) * lama_sewa
                for qty, harga in zip(qty_list, harga_list)
            )

            # Update transaksi utama
            cursor.execute("""
                UPDATE transactions
                SET customer_id=%s, user_id=%s, no_nota=%s, tanggal_nota=%s,
                    tanggal_sewa=%s, tanggal_kembali=%s, lama_sewa=%s,
                    status_pembayaran=%s, status_pengembalian=%s, jaminan=%s, note=%s, total=%s
                WHERE transaction_id=%s
            """, (
                customer_id, user_id, no_nota, tanggal_nota,
                tanggal_sewa, tanggal_kembali, lama_sewa,
                status_pembayaran, status_penegembalian, jaminan, note, total, transaction_id
            ))

            # Hapus detail lama
            cursor.execute("DELETE FROM transaction_details WHERE transaction_id = %s", (transaction_id,))

            # Simpan ke tabel transaction_details
            product_ids = request.form.getlist("product_id[]")
            for product_id, qty, harga in zip(product_ids, qty_list, harga_list):
                cursor.execute("""
                    INSERT INTO transaction_details (transaction_id, product_id, product_qty, harga_sewa)
                    VALUES (%s, %s, %s, %s)
                """, (transaction_id, product_id, qty, harga))

            conn.commit()
            conn.close()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({
                    'status': 'success',
                    'message': 'Transaksi berhasil diperbarui.',
                    'redirect_url': url_for('detail_transaksi', transaction_id=transaction_id)})
            else:
                flash("Transaksi berhasil diperbarui.", "success")
                return redirect(url_for('detail_transaksi', transaction_id=transaction_id))

        except Exception as e:
            conn.rollback()
            print(f"[ERROR] Gagal memperbarui transaksi: {e}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'status': 'error', 'message': 'Terjadi kesalahan saat menyimpan perubahan.'}), 500
            else:
                flash("Terjadi kesalahan saat menyimpan perubahan.", "danger")
                abort(500)

        finally:
            cursor.close()
            conn.close()

    # === GET METHOD ===
    try:
        # Ambil data transaksi
        cursor.execute("""
            SELECT t.*,
                       u.name as user_name,
                       c.name as customer_name
                        FROM transactions t 
            JOIN customers c ON t.customer_id = c.customer_id
            JOIN users u ON t.user_id = u.user_id
            WHERE t.transaction_id = %s
        """, (transaction_id,))
        transaction = cursor.fetchone()
        if not transaction:
            abort(404)

        # Ambil detail produk
        cursor.execute("""
            SELECT td.*, p.name, p.qty as stock, p.harga_sewa,
                       t.lama_sewa
            FROM transaction_details td
            JOIN products p ON td.product_id = p.product_id
            JOIN transactions t ON td.transaction_id = t.transaction_id
            WHERE td.transaction_id = %s
        """, (transaction_id,))
        product_details = cursor.fetchall()

        # Ambil semua produk (opsional, jika ingin user bisa ganti produk)
        cursor.execute("""
            SELECT 
                product_id, 
                name, 
                qty as stock, 
                harga_sewa 
            FROM products
        """)
        all_products = cursor.fetchall()

        # Ambil semua customer (untuk dropdown)
        cursor.execute("SELECT customer_id, name FROM customers")
        customers = cursor.fetchall()

        return render_template('transactions/edit_transaksi.html',
                               transaction=transaction,
                               product_details=product_details,
                               all_products=all_products,
                               customers=customers,
                               user=current_user)

    except Exception as e:
        print(f"[ERROR]: {e}")
        traceback.print_exc()
        abort(500)

#--- Route: Hapus transaksi --- 
@app.route('/delete_transaksi/<int:transaction_id>', methods=['GET', 'POST'])
@login_required
def delete_transaksi(transaction_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Hapus detail transaksi terlebih dahulu
        cursor.execute("DELETE FROM transaction_details WHERE transaction_id = %s", (transaction_id,))

        # Hapus transaksi utama
        cursor.execute("DELETE FROM transactions WHERE transaction_id = %s", (transaction_id,))

        conn.commit()
        flash("Transaksi berhasil dihapus.", "success")
        return redirect(url_for('riwayat_transaksi'))  # ganti 'list_transaksi' sesuai nama route daftar transaksi

    except Exception as e:
        conn.rollback()
        print(f"[ERROR] Gagal menghapus transaksi: {e}")
        flash("Terjadi kesalahan saat menghapus transaksi.", "danger")
        abort(500)

    finally:
        cursor.close()
        conn.close()
    

with open("static/img/logo.png", "rb") as image_file:
    logo_base64 = base64.b64encode(image_file.read()).decode("utf-8")

@app.route("/cetak-nota/<int:transaction_id>")
def cetak_nota(transaction_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Ambil transaksi + customer + user
        cursor.execute("""
            SELECT t.*, 
                   c.name AS customer_name, c.phone_number AS customer_phone, c.address AS customer_address,
                   u.name AS user_name
            FROM transactions t
            JOIN customers c ON t.customer_id = c.customer_id
            JOIN users u ON t.user_id = u.user_id
            WHERE t.transaction_id = %s
        """, (transaction_id,))
        transaction = cursor.fetchone()

        if not transaction:
            abort(404, "Transaksi tidak ditemukan.")

        # Ambil detail produk
        cursor.execute("""
            SELECT td.product_qty AS qty, td.harga_sewa AS harga, p.name, p.product_id, t.lama_sewa
            FROM transaction_details td
            JOIN products p ON td.product_id = p.product_id
            JOIN transactions t ON td.transaction_id = t.transaction_id
            WHERE td.transaction_id = %s
        """, (transaction_id,))
        details = cursor.fetchall()

        cursor.close()
        conn.close()

        # Render HTML
        rendered = render_template("transactions/nota_pdf.html", data={
            'transaction': transaction,
            'details': details,
        }, now=datetime.now(), logo_base64=logo_base64)

        # Buat PDF sementara
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as pdf_file:
            options = {
                'page-size': 'B5',
                'margin-top': '10mm',
                'margin-right': '10mm',
                'margin-bottom': '10mm',
                'margin-left': '10mm',
                'encoding': "UTF-8",
            }
            pdfkit.from_string(rendered, pdf_file.name, configuration=pdfkit_config, options=options)

            pdf_file_path = pdf_file.name

        return send_file(pdf_file_path, mimetype='application/pdf')

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}", 500


## MODULE LAPORAN
#--- Route: Halaman laporan transaksi ---
@app.route('/laporan', methods=['GET', 'POST'])
@login_required
def laporan():
    tanggal_dari = request.args.get('dari')
    tanggal_sampai = request.args.get('sampai')
    status_pembayaran = request.args.get('status_pembayaran') 
    status_pengembalian = request.args.get('status_pengembalian')

    data = []
    ringkasan = {
        "total_transaksi": 0,
        "total_produk": 0,
        "total_belum_lunas": 0,
        "total_belum_kembali": 0,
        "total_pendapatan": 0
    }

    if tanggal_dari and tanggal_sampai:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Ambil data transaksi dan detail produk
        query = """
            SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.lama_sewa,
                   t.tanggal_kembali, t.status_pembayaran, t.status_pengembalian, c.name AS customer_name,
                   u.name AS user_name, p.name AS product_name, td.product_qty, p.harga_sewa,
                   (td.product_qty * p.harga_sewa * t.lama_sewa) AS total_harga
            FROM transactions t
            JOIN users u ON t.user_id = u.user_id
            JOIN customers c ON t.customer_id = c.customer_id
            JOIN transaction_details td ON t.transaction_id = td.transaction_id
            JOIN products p ON td.product_id = p.product_id
            WHERE t.tanggal_nota BETWEEN %s AND %s 
        """

        params = [tanggal_dari, tanggal_sampai]

        # Tambahkan filter status pembayaran
        if status_pembayaran in ['lunas', 'belum']:
            query += " AND t.status_pembayaran = %s"
            status_pembayaran_val = 'lunas' if status_pembayaran == 'lunas' else 'belum lunas'
            params.append(status_pembayaran_val)

        # Tambahkan filter status pengembalian
        if status_pengembalian in ['kembali', 'belum']:
            query += " AND t.status_pengembalian = %s"
            status_pengembalian_val = 'sudah' if status_pengembalian == 'kembali' else 'belum'
            params.append(status_pengembalian_val)

        query += " ORDER BY t.tanggal_nota DESC"

        cursor.execute(query, params)
        data = cursor.fetchall()

        # Ringkasan
        ringkasan["total_transaksi"] = len(set(row['transaction_id'] for row in data))
        ringkasan["total_produk"] = sum(row['product_qty'] for row in data)
        ringkasan["total_belum_lunas"] = sum(row['product_qty'] for row in data if row['status_pembayaran'].lower() != 'lunas')
        ringkasan["total_belum_kembali"] = sum(row['product_qty'] for row in data if row['status_pengembalian'].lower() != 'kembali')
        ringkasan["total_pendapatan"] = sum(row['total_harga'] or 0 for row in data)

        cursor.close()
        conn.close()

    return render_template(
        'report/laporan.html',
        data=data,
        tanggal_dari=tanggal_dari,
        tanggal_sampai=tanggal_sampai,
        status_pembayaran=status_pembayaran,
        status_pengembalian=status_pengembalian,
        ringkasan=ringkasan,
        user=current_user
    )

#--- Route: Cetak PDF laporan transaksi ---
@app.route("/laporan/pdf")
def export_pdf():
    try:
        tanggal_dari = request.args.get('dari')
        tanggal_sampai = request.args.get('sampai')
        status_pembayaran = request.args.get('status_pembayaran') 
        status_pengembalian = request.args.get('status_pengembalian')
    
        data = []
        ringkasan = {
            "total_transaksi": 0,
            "total_produk": 0,
            "total_belum_lunas": 0,
            "total_belum_kembali": 0,
            "total_pendapatan": 0
        }
    
        if tanggal_dari and tanggal_sampai:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
    
            # Ambil data transaksi dan detail produk
            query = """
                SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.lama_sewa,
                       t.tanggal_kembali, t.status_pembayaran, t.status_pengembalian, c.name AS customer_name,
                       u.name AS user_name, p.name AS product_name, td.product_qty, p.harga_sewa,
                       (td.product_qty * p.harga_sewa * t.lama_sewa) AS total_harga
                FROM transactions t
                JOIN users u ON t.user_id = u.user_id
                JOIN customers c ON t.customer_id = c.customer_id
                JOIN transaction_details td ON t.transaction_id = td.transaction_id
                JOIN products p ON td.product_id = p.product_id
                WHERE t.tanggal_nota BETWEEN %s AND %s 
            """
    
            params = [tanggal_dari, tanggal_sampai]
    
            # Tambahkan filter status pembayaran
            if status_pembayaran in ['lunas', 'belum']:
                query += " AND t.status_pembayaran = %s"
                status_pembayaran_val = 'lunas' if status_pembayaran == 'lunas' else 'belum lunas'
                params.append(status_pembayaran_val)
    
            # Tambahkan filter status pengembalian
            if status_pengembalian in ['kembali', 'belum']:
                query += " AND t.status_pengembalian = %s"
                status_pengembalian_val = 'sudah' if status_pengembalian == 'kembali' else 'belum'
                params.append(status_pengembalian_val)
    
            query += " ORDER BY t.tanggal_nota DESC"
    
            cursor.execute(query, params)
            data = cursor.fetchall()
    
            # Ringkasan
            ringkasan["total_transaksi"] = len(set(row['transaction_id'] for row in data))
            ringkasan["total_produk"] = sum(row['product_qty'] for row in data)
            ringkasan["total_belum_lunas"] = sum(row['product_qty'] for row in data if row['status_pembayaran'].lower() != 'lunas')
            ringkasan["total_belum_kembali"] = sum(row['product_qty'] for row in data if row['status_pengembalian'].lower() != 'kembali')
            ringkasan["total_pendapatan"] = sum(row['total_harga'] or 0 for row in data)
    
            cursor.close()
            conn.close()

        # Render HTML
        rendered = render_template('report/laporan_pdf.html', data=data, ringkasan=ringkasan,
                               tanggal_dari=tanggal_dari, tanggal_sampai=tanggal_sampai,
                               now=datetime.now(), logo_base64=logo_base64, user=current_user)

        # Buat PDF sementara
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as pdf_file:
            options = {
                'page-size': 'A4',
                'orientation': 'Landscape',
                'margin-top': '15mm',
                'margin-bottom': '15mm',
                'margin-left': '15mm',
                'margin-right': '15mm',
                'encoding': 'UTF-8',
            }

            pdfkit.from_string(rendered, pdf_file.name, configuration=pdfkit_config, options=options)

            pdf_file_path = pdf_file.name

        return send_file(pdf_file_path, mimetype='application/pdf')

    except Exception as e:
        return f"Terjadi kesalahan: {str(e)}", 500

#--- Route: Export laporan transaksi ke Excel ---
@app.route('/laporan/excel')
@login_required
def export_excel():
    tanggal_dari = request.args.get('dari')
    tanggal_sampai = request.args.get('sampai')
    status_pembayaran = request.args.get('status_pembayaran') 
    status_pengembalian = request.args.get('status_pengembalian')

    data = []
    ringkasan = {
        "total_transaksi": 0,
        "total_produk": 0,
        "total_belum_lunas": 0,
        "total_belum_kembali": 0,
        "total_pendapatan": 0
    }

    if tanggal_dari and tanggal_sampai:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
            SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.lama_sewa,
                   t.tanggal_kembali, t.status_pembayaran, t.status_pengembalian, c.name AS customer_name,
                   u.name AS user_name, p.name AS product_name, td.product_qty, p.harga_sewa,
                   (td.product_qty * p.harga_sewa * t.lama_sewa) AS total_harga
            FROM transactions t
            JOIN users u ON t.user_id = u.user_id
            JOIN customers c ON t.customer_id = c.customer_id
            JOIN transaction_details td ON t.transaction_id = td.transaction_id
            JOIN products p ON td.product_id = p.product_id
            WHERE t.tanggal_nota BETWEEN %s AND %s 
        """

        params = [tanggal_dari, tanggal_sampai]

        # Tambahkan filter status pembayaran
        if status_pembayaran in ['lunas', 'belum']:
            query += " AND t.status_pembayaran = %s"
            status_pembayaran_val = 'lunas' if status_pembayaran == 'lunas' else 'belum lunas'
            params.append(status_pembayaran_val)

        # Tambahkan filter status pengembalian
        if status_pengembalian in ['kembali', 'belum']:
            query += " AND t.status_pengembalian = %s"
            status_pengembalian_val = 'sudah' if status_pengembalian == 'kembali' else 'belum'
            params.append(status_pengembalian_val)

        query += " ORDER BY t.tanggal_nota DESC"

        cursor.execute(query, params)
        data = cursor.fetchall()

        # Ringkasan
        ringkasan["total_transaksi"] = len(set(row['transaction_id'] for row in data))
        ringkasan["total_produk"] = sum(row['product_qty'] for row in data)
        ringkasan["total_belum_lunas"] = sum(row['product_qty'] for row in data if row['status_pembayaran'].lower() != 'lunas')
        ringkasan["total_belum_kembali"] = sum(row['product_qty'] for row in data if row['status_pengembalian'].lower() != 'kembali')
        ringkasan["total_pendapatan"] = sum(row['total_harga'] or 0 for row in data)

        cursor.close()
        conn.close()

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi Persewaan"

    # Gaya
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    currency_format = '#,##0'
    border_style = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    # Header
    headers = [
        'No Nota', 'Tanggal Nota', 'Tanggal Sewa', 'Tanggal Kembali', 'Pelanggan', 'Produk', 'Qty', 'Lama Sewa', 'Harga Sewa', 
        'Subtotal', 'Pembayaran', 'Pengembalian', 'Petugas'
    ]
    ws.append(headers)

    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = border_style

    # Data
    for i, row in enumerate(data, start=2):
        ws.append([
            row['no_nota'],
            row['tanggal_nota'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['tanggal_sewa'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['tanggal_kembali'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['customer_name'],
            row['product_name'],
            row['product_qty'],
            row['lama_sewa'],
            row['harga_sewa'],
            row['total_harga'],
            row['status_pembayaran'],
            row['status_pengembalian'],
            row['user_name']
        ])
        for col in range(1, 14):
            cell = ws.cell(row=i, column=col)
            cell.border = border_style
            if col in [9, 10]:  # Harga sewa dan total
                cell.number_format = currency_format
    
    # Auto-width kolom
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Spasi
    ws.append([])
    ws.append(['Ringkasan'])

    # Ringkasan
    ring_start_row = ws.max_row + 1
    for label, key in [
        ('Total Transaksi', 'total_transaksi'),
        ('Total Produk Disewa', 'total_produk'),
        ('Belum Lunas', 'total_belum_lunas'),
        ('Belum Kembali', 'total_belum_kembali'),
        ('Total Pendapatan', 'total_pendapatan')
    ]:
        ws.append([label, ringkasan[key]])

    # Styling ringkasan
    for i in range(ring_start_row, ws.max_row + 1):
        for col in range(1, 3):
            cell = ws.cell(row=i, column=col)
            cell.border = border_style
            if col == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = currency_format

    # Simpan ke memori
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nama_file = f"Laporan_Transaksi_{tanggal_dari}_sd_{tanggal_sampai}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nama_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

#--- Route: Export laporan transaksi ke CSV ---
@app.route('/laporan/csv')
def export_csv():
    tanggal_dari = request.args.get('dari')
    tanggal_sampai = request.args.get('sampai')
    status_pembayaran = request.args.get('status_pembayaran')
    status_pengembalian = request.args.get('status_pengembalian')

    if not tanggal_dari or not tanggal_sampai:
        return "Tanggal tidak lengkap", 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT t.transaction_id, t.no_nota, t.tanggal_nota, t.tanggal_sewa, t.lama_sewa,
               t.tanggal_kembali, t.status_pembayaran, t.status_pengembalian, c.name AS customer_name,
               u.name AS user_name, p.name AS product_name, td.product_qty, p.harga_sewa,
               (td.product_qty * p.harga_sewa * t.lama_sewa) AS total_harga
        FROM transactions t
        JOIN users u ON t.user_id = u.user_id
        JOIN customers c ON t.customer_id = c.customer_id
        JOIN transaction_details td ON t.transaction_id = td.transaction_id
        JOIN products p ON td.product_id = p.product_id
        WHERE t.tanggal_nota BETWEEN %s AND %s
    """
    params = [tanggal_dari, tanggal_sampai]

    if status_pembayaran in ['lunas', 'belum']:
        query += " AND t.status_pembayaran = %s"
        params.append('lunas' if status_pembayaran == 'lunas' else 'belum lunas')

    if status_pengembalian in ['kembali', 'belum']:
        query += " AND t.status_pengembalian = %s"
        params.append('sudah' if status_pengembalian == 'kembali' else 'belum')

    query += " ORDER BY t.tanggal_nota DESC"
    cursor.execute(query, params)
    data = cursor.fetchall()

    # Hitung ringkasan
    ringkasan = {
        "total_transaksi": len(set(row['transaction_id'] for row in data)),
        "total_produk": sum(row['product_qty'] for row in data),
        "total_belum_lunas": sum(row['product_qty'] for row in data if row['status_pembayaran'].lower() != 'lunas'),
        "total_belum_kembali": sum(row['product_qty'] for row in data if row['status_pengembalian'].lower() != 'kembali'),
        "total_pendapatan": sum(row['total_harga'] or 0 for row in data)
    }

    cursor.close()
    conn.close()

    # Tulis CSV
    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Header
    writer.writerow([
        'No Nota', 'Tanggal Nota', 'Tanggal Sewa', 'Tanggal Kembali', 'Pelanggan', 'Produk', 'Qty', 'Lama Sewa', 'Harga Sewa', 
        'Subtotal', 'Pembayaran', 'Pengembalian', 'Petugas'
    ])

    # Data
    for row in data:
        writer.writerow([
            row['no_nota'],
            row['tanggal_nota'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['tanggal_sewa'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['tanggal_kembali'].strftime('%d-%m-%Y') if row['tanggal_nota'] else '',
            row['customer_name'],
            row['product_name'],
            row['product_qty'],
            row['lama_sewa'],
            row['harga_sewa'],
            row['total_harga'],
            row['status_pembayaran'],
            row['status_pengembalian'],
            row['user_name']
        ])

    # Spasi dan Ringkasan
    writer.writerow([])
    writer.writerow(['Ringkasan'])
    writer.writerow(['Total Transaksi', ringkasan["total_transaksi"]])
    writer.writerow(['Total Produk Disewa', ringkasan["total_produk"]])
    writer.writerow(['Belum Lunas', ringkasan["total_belum_lunas"]])
    writer.writerow(['Belum Kembali', ringkasan["total_belum_kembali"]])
    writer.writerow(['Total Pendapatan', ringkasan["total_pendapatan"]])

    # Response
    filename = f"Laporan_Transaksi_{tanggal_dari}_sd_{tanggal_sampai}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



if __name__ == '__main__':
    app.run(debug=True)